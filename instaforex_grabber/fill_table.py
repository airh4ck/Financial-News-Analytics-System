import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from os.path import exists
import json
import re

from KafkaProducer import uploadData


def find_element_if_exists(driver, by, name):
    try:
        return driver.find_element(by, name)
    except NoSuchElementException:
        return None


def transform(obj, f):
    return None if obj is None else f(obj)


if __name__ == '__main__':
    # table_name = sys.argv[1]
    # authors_dump_name = sys.argv[2]
    currency = sys.argv[1]

    # wb = openpyxl.load_workbook(filename = table_name) if exists(table_name) else openpyxl.Workbook()
    # ws = wb[currency] if currency in wb.sheetnames else wb.create_sheet(currency)

    # authors_struct = {'last' : 0, 'authors' : {}}
    # if exists(authors_dump_name):
    #     authors_dump = open(authors_dump_name, 'r')
    #     authors_struct = json.loads(authors_dump.read())
    #     authors_dump.close()
    # new_id = authors_struct['last']
    # authors = authors_struct['authors']

    opts = webdriver.ChromeOptions()
    opts.headless = True
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])

    print("Connecting to driver...")
    driver = webdriver.Remote(
        "http://selenium:4444/wd/hub", options=opts)
    print("Successfully connected!")

    wd_wait = WebDriverWait(driver, 10)

    url = input()
    while url:
        print(url)

        try:
            driver.get(url)
            wd_wait.until(EC.all_of(
                EC.presence_of_element_located((By.ID, 'article-title')),
                EC.presence_of_element_located(
                    (By.CLASS_NAME, 'stat_view_metric')),
                EC.presence_of_element_located((By.CLASS_NAME, 'ioa_text')),
                EC.presence_of_element_located((By.CLASS_NAME, 'time_author'))
            ))

            title = driver.find_element(By.ID, 'article-title').text
            main_part = driver.find_element(By.CLASS_NAME, 'stat_view_metric')
            paragraphs = main_part.find_elements(By.TAG_NAME, 'p')
            body = '\n'.join(map(lambda p: p.text, paragraphs))
            author = driver.find_element(
                By.CLASS_NAME, 'ioa_text').text.split('\n')[2]

            date = re.search(
                r'(0[1-9]|[12][0-9]|3[01])\.(0[1-9]|1[012])\.(19|20)\d\d',
                driver.find_element(By.CLASS_NAME, 'time_author')
                .find_element(By.CLASS_NAME, 'time_span')
                .text
            ).group(0)

            time = re.search(
                r'([0-1]?[0-9]|2[0-3]):[0-5][0-9]',
                driver.find_element(By.CLASS_NAME, 'time_author')
                .find_element(By.CLASS_NAME, 'time_span')
                .text
            ).group(0)

            time = driver.find_element(By.CLASS_NAME, 'time').text

            # author_id = new_id
            # if author in authors:
            #     author_id = authors[author]
            # else:
            #     authors[author] = author_id
            #     new_id += 1

            # uploader.update_row({
            #     'pair': currency,
            #     'title': title,
            #     'author': author,
            #     'link': url,
            #     'text': body,
            #     'date': date,
            #     'time': time
            # })
            uploadData({
                'pair': currency,
                'title': title,
                'author': author,
                'link': url,
                'text': body,
                'date': date,
                'time': time
            })

            print('+')

        except Exception as e:
            print(e)

        finally:
            url = input()

    print()
    driver.quit()
